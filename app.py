import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from engine import build_report

# --------- ДОБАВЛЕНО ДЛЯ GOOGLE SHEETS ---------
import gspread
from google.oauth2.service_account import Credentials

# ID твоей таблицы (кусок из URL между /d/ и /edit)
SHEET_ID = "12NIk4vQ0Z7av6b4JbAIVKyY_blYnb5Vacumy_4FCTdM"

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Используем локальный файл gcp_service_key.json из репозитория
creds = Credentials.from_service_account_file(
    "gcp_service_key.json",
    scopes=SCOPES,
)

try:
    gs_client = gspread.authorize(creds)
    sheet = gs_client.open_by_key(SHEET_ID).sheet1  # первый лист таблицы
except Exception as e:
    st.error("Ошибка при подключении к Google Sheets:")
    st.code(repr(e))
    st.stop()
# --------- /ДОБАВЛЕНО ДЛЯ GOOGLE SHEETS ---------

def register_client_run(client_id: str, max_free_runs: int = 1):
    """
    Регистрирует запуск клиента в Google Sheets.
    Возвращает (allowed: bool, free_left: int).

    max_free_runs — сколько бесплатных запусков даём новому клиенту.
    """
    # читаем все строки как словари
    records = sheet.get_all_records()  # [{'client_id': ..., 'free_runs_left': ...}, ...]

    # ищем клиента в уже существующих строках
    for idx, row in enumerate(records, start=2):  # данные начинаются со 2-й строки (1 — заголовки)
        if row.get("client_id") == client_id:
            free_left = int(row.get("free_runs_left") or 0)
            total_runs = int(row.get("total_runs") or 0)

            # если бесплатных запусков не осталось — блокируем
            if free_left <= 0:
                return False, free_left

            # уменьшаем оставшиеся, увеличиваем общее число запусков
            free_left -= 1
            total_runs += 1

            # обновляем ячейки в таблице
            sheet.update_cell(idx, 2, free_left)  # колонка B: free_runs_left
            sheet.update_cell(idx, 3, total_runs)  # колонка C: total_runs
            sheet.update_cell(idx, 4, pd.Timestamp.utcnow().isoformat())  # колонка D: last_run

            return True, free_left

    # если клиента не нашли — создаём новую строку
    free_left = max_free_runs - 1
    total_runs = 1
    sheet.append_row(
        [
            client_id,
            free_left,
            total_runs,
            pd.Timestamp.utcnow().isoformat(),
        ]
    )

    return True, free_left
# --------- /ДОБАВЛЕНО ДЛЯ GOOGLE SHEETS ---------

# ---------------- НАСТРОЙКИ СТРАНИЦЫ ----------------
st.set_page_config(
    page_title="Умный отчет",
    page_icon="📊",
    layout="wide",
)

# ---------------- ГЛОБАЛЬНЫЙ СТИЛЬ (CSS) ----------------
st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(135deg, #e4f0ff 0%, #ffffff 55%) !important;
        color: #102A43 !important;
        font-size: 16px !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important;
    }

    .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
    }

    /* === ЗАГРУЗЧИК ФАЙЛОВ === */

    [data-testid="stFileUploader"] section {
        background-color: #f5f7fb !important;
        border: 1px solid #d0d7ea !important;
        border-radius: 8px !important;
        color: #102A43 !important;
    }

    [data-testid="stFileDropzone"] span,
    [data-testid="stFileUploaderInstructions"] {
        display: none !important;
    }

    [data-testid="stFileUploader"] button {
        background-color: #eef3ff !important;
        color: #003366 !important;
        border: 1px solid #d0d7ea !important;
        border-radius: 6px !important;
        padding: 6px 14px !important;
        font-weight: 600 !important;
        box-shadow: none !important;
    }
    [data-testid="stFileUploader"] button:hover {
        background-color: #d6e4ff !important;
    }

    [data-testid="stFileDropzone"] {
        background-color: transparent !important;
        border: none !important;
    }

    [data-testid="stFileUploaderFileName"] {
        color: #003366 !important;
        background-color: #ffffff !important;
        padding: 4px 8px !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        display: inline-block !important;
    }
    [data-testid="stFileUploaderSize"] {
        color: #4a637e !important;
        background-color: #ffffff !important;
        padding: 2px 6px !important;
        border-radius: 4px !important;
        margin-left: 4px !important;
        font-size: 13px !important;
    }

    .stButton > button, .stDownloadButton > button {
        background-color: #1E88E5 !important;
        color: white !important;
        border-radius: 8px !important;
        padding: 10px 22px !important;
        font-size: 16px !important;
        border: none !important;
        font-weight: 600 !important;
        transition: 0.3s ease-in-out;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #1565C0 !important;
        transform: translateY(-1px);
    }

    h1, h2, h3, h4 {
        color: #102A43 !important;
        font-weight: 700 !important;
    }

    .file-label {
        padding: 4px 10px;
        margin: 4px 0;
        border-radius: 6px;
        background-color: #eef3ff;
        color: #003366;
        font-weight: 600;
        display: inline-block;
    }

    [data-testid="stDataFrame"] div[role="grid"] {
        background-color: #ffffff !important;
        color: #102A43 !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ---------------- ГЛАВНЫЙ ЗАГОЛОВОК ----------------
st.markdown(
    """
    <div style="text-align: center; padding: 20px; background-color: #F0F4FF;
                border-radius: 10px; margin-bottom: 1.5rem;">
        <h2 style="color: #003366; margin-bottom: 0.5rem;">
            📊 Умный контроль рабочего времени
        </h2>
        <p style="color: #003366; font-size:16px; margin: 0;">
            Загрузите журнал проходов и (по желанию) файл кадров — система автоматически сформирует табель,
            рассчитает недоработки, выходы, длительные отсутствия и причины прогула.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------- ПРИМЕРЫ ФАЙЛОВ ----------------
import base64
import os

st.header("📂 Пример загружаемых файлов")

def download_file(path, label):
    with open(path, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(path)}">{label}</a>'
    st.markdown(href, unsafe_allow_html=True)

col_example1, col_example2 = st.columns(2)

with col_example1:
    download_file("examples/пример СКУД.xlsx", "⬇ Скачать пример отчёта пропусков (СКУД)")

with col_example2:
    download_file("examples/пример от кадров.xlsx", "⬇ Скачать пример кадрового файла")

st.markdown("---")

# --- Шаг 1. Загрузка файлов ---
st.header("Шаг 1. Загрузка файлов")

col_left, col_right = st.columns([2, 1])

with col_left:
    # -------- ЖУРНАЛ ПРОХОДОВ --------
    st.subheader("📘 Журнал проходов")

    st.markdown(
        """
        <div style="
            padding: 10px; 
            background-color: #eef3ff; 
            border-radius: 6px; 
            border: 1px solid #d0d7ea; 
            margin-bottom: 8px; 
            color:#003366;
        ">
            📤 <b>Загрузите файл журнала проходов</b><br>
            <span style="font-size: 14px;">
                Формат: XLS или XLSX, размер до 200 МБ.
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    file_journal = st.file_uploader(
        "Журнал проходов",
        type=["xls", "xlsx"],
        label_visibility="collapsed",
        help="Файл журнала из системы проходов (XLS/XLSX).",
    )

    st.markdown("---")

    # -------- ФАЙЛ КАДРОВ --------
    st.subheader("📗 Сведения из кадров (по желанию)")

    st.markdown(
        """
        <div style="
            padding: 10px; 
            background-color: #eef3ff; 
            border-radius: 6px; 
            border: 1px solid #d0d7ea; 
            margin-bottom: 8px; 
            color:#003366;
        ">
            📤 <b>Загрузите файл с отсутствиями (кадровый отчёт)</b><br>
            <span style="font-size: 14px;">
                Отпуска, больничные, командировки и др. причины отсутствий. 
                Можно не загружать — тогда колонка «Причина отсутствия» останется пустой.
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    kadry_file = st.file_uploader(
        "Загрузите файл кадров (.xls / .xlsx)",
        type=["xls", "xlsx"],
    )

with col_right:
    st.markdown(
        """
        **Подсказки:**
        - Журнал — стандартная выгрузка из системы проходов.
        - Кадровый файл — со столбцами:
          *«Сотрудник», «Вид отсутствия», «с», «до»*.
        - Можно загрузить только журнал —
          тогда «Причина отсутствия» останется пустой.
        """,
        unsafe_allow_html=False,
    )

# Если журнал не загружен — дальше не идём
if file_journal is None:
    st.warning("⬆ Сначала загрузите файл журнала проходов.")
    st.stop()

st.caption("Перетащите файл сюда или нажмите «Browse files» для выбора файла журнала.")

# Пояснение по кадровому файлу
if kadry_file is None:
    st.info(
        "Кадровый файл *не обязателен*. "
        "Можете загрузить его для указания причин отсутствия "
        "или сразу перейти к обработке."
    )
else:
    st.success("✅ Оба файла загружены!")

# Красивый вывод названий файлов
st.markdown(
    f"<div class='file-label'>📘 Журнал: {file_journal.name}</div>",
    unsafe_allow_html=True,
)
if kadry_file is not None:
    st.markdown(
        f"<div class='file-label'>📗 Кадровый файл: {kadry_file.name}</div>",
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        "<div class='file-label' style='background-color:#f5f5f5; color:#555;'>"
        "📗 Кадровый файл: не загружен"
        "</div>",
        unsafe_allow_html=True,
    )

# ---------------- ШАГ 2. ОБРАБОТКА ДАННЫХ ----------------
st.header("Шаг 2. Обработка данных")
# Идентификатор клиента (email или Telegram)
st.subheader("Кто запускает отчёт?")

client_id = st.text_input(
    "Укажите ваш email или Telegram (@username)",
    help="Это нужно, чтобы дать вам 1 бесплатный запуск и учитывать дальнейшие обращения.",
)

st.caption(
    "Мы не рассылаем спам. Идентификатор используется только для учёта запусков и поддержки."
)

final_df = None

if st.button("🚀 Обработать данные"):
    # 1. Проверяем, что клиент указал идентификатор
    clean_client_id = (client_id or "").strip()
    if not clean_client_id:
        st.warning("Сначала укажите ваш email или Telegram выше.")
    else:
        # 2. Пытаемся зарегистрировать запуск в Google Sheets
        try:
            allowed, free_left = register_client_run(clean_client_id)
        except Exception as e:
            st.error(f"❌ Не удалось связаться с системой учёта запусков: {e}")
        else:
            if not allowed:
                st.error(
                    "😔 Похоже, бесплатные запуски для этого идентификатора закончились.\n\n"
                    "Напишите, пожалуйста, автору сервиса, чтобы подключить платный доступ "
                    "или выдать дополнительные тестовые запуски."
                )
            else:
                # 3. Разрешено — запускаем обработку
                try:
                    final_df = build_report(file_journal, kadry_file)
                except Exception as e:
                    st.error(f"❌ Ошибка при обработке данных: {e}")
                else:
                    st.success(
                        f"✅ Обработка завершена. "
                        f"Осталось бесплатных запусков: {free_left}."
                    )

# Если ещё не нажали кнопку или произошла ошибка — дальше не идём
if final_df is None:
    st.stop()

# ---------------- ШАГ 3. ПРЕДПРОСМОТР И ВЫГРУЗКА ----------------
st.header("Шаг 3. Выгрузка отчёта")

# Базовый набор колонок
visible_cols = [
    "ФИО",
    "Дата",
    "Время прихода",
    "Время ухода",
    "Опоздание",
    "Общее время",
    "Вне офиса",
    "Выходы",
    "Отсутствие более 2 часов подряд",
    "Итого за день",
    "Итого за неделю",
    "Недоработки",
    "Причина отсутствия",
]

visible_cols = [c for c in visible_cols if c in final_df.columns]

if not visible_cols:
    st.warning("В итоговом отчёте нет ожидаемых колонок для отображения.")
    final_view = final_df.copy()
else:
    final_view = final_df[visible_cols].copy()

# Сортировка по ФИО и дате (если возможно)
if "ФИО" in final_view.columns and "Дата" in final_view.columns:
    final_view = final_view.sort_values(["ФИО", "Дата"])

# ---------------- ФОРМИРОВАНИЕ И СКАЧИВАНИЕ EXCEL ----------------
buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    sheet_name = "Журнал"

    # пишем таблицу с отступом (чтобы сверху уместить заголовок)
    final_view.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)

    wb = writer.book
    ws = writer.sheets[sheet_name]

    max_col = ws.max_column
    last_col_letter = get_column_letter(max_col)

    # --- Большой заголовок ---
    title_cell = ws["A1"]
    title_cell.value = "ОТЧЁТ ЗА НЕДЕЛЮ"
    title_cell.font = Font(name="Times New Roman", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A1:{last_col_letter}1")

    # --- Шапка таблицы (строка 4) ---
    header_row = 4
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(name="Times New Roman", size=11, bold=True)

    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # --- Выравнивание данных и ширина столбцов ---
    col_names = [cell.value for cell in ws[header_row]]

    for col_idx, name in enumerate(col_names, start=1):
        align = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        for row in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = align

    width_map = {
        "ФИО": 30,
        "Дата": 12,
        "Время прихода": 15,
        "Время ухода": 15,
        "Опоздание": 14,
        "Общее время": 14,
        "Вне офиса": 16,
        "Выходы": 12,
        "Отсутствие более 2 часов подряд": 28,
        "Итого за день": 14,
        "Итого за неделю": 16,
        "Недоработки": 16,
        "Причина отсутствия": 28,
    }

    for col_idx, name in enumerate(col_names, start=1):
        if name in width_map:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width_map[name]

    base_font = Font(name="Times New Roman", size=11)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = base_font

    ws.freeze_panes = "A5"

buffer.seek(0)

st.download_button(
    label="💾 Скачать итоговый отчёт (Excel)",
    data=buffer,
    file_name="умный_табель.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)










